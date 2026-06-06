"""Excel export for generated games and check results."""
from __future__ import annotations

import os
from datetime import datetime
from typing import List, Optional

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from models.lottery_types import (
    LotteryType, LOTTERY_CONFIGS,
    TIMEMANIA_TEAMS, DIA_DE_SORTE_MONTHS,
)
from models.ticket import Ticket, DrawResult
from services.checker import CheckResult


# ── Lottery brand colours (hex, no #) ────────────────────────────────────────
BRAND: dict[LotteryType, str] = {
    LotteryType.MEGA_SENA:       "1A7F44",
    LotteryType.QUINA:           "1A478C",
    LotteryType.LOTOFACIL:       "8B2FC9",
    LotteryType.LOTOMANIA:       "E35208",
    LotteryType.TIMEMANIA:       "1B6E35",
    LotteryType.DUPLA_SENA:      "B71C1C",
    LotteryType.DIA_DE_SORTE:    "D4820A",
    LotteryType.SUPER_SETE:      "00838F",
    LotteryType.MAIS_MILIONARIA: "6A0DAD",
}

# Lighter tint for number cells (first 6 hex chars = colour)
LIGHT: dict[LotteryType, str] = {
    LotteryType.MEGA_SENA:       "D5F5E3",
    LotteryType.QUINA:           "D6EAF8",
    LotteryType.LOTOFACIL:       "EDE7F6",
    LotteryType.LOTOMANIA:       "FDEBD0",
    LotteryType.TIMEMANIA:       "D5F5E3",
    LotteryType.DUPLA_SENA:      "FFEBEE",
    LotteryType.DIA_DE_SORTE:    "FFF8E1",
    LotteryType.SUPER_SETE:      "E0F7FA",
    LotteryType.MAIS_MILIONARIA: "F3E5F5",
}

MATCH_FILL  = PatternFill("solid", fgColor="27AE60")   # matched number
MATCH_FONT  = Font(bold=True, color="FFFFFF")
NO_MATCH    = PatternFill("solid", fgColor="ECEFF1")   # drawn but not picked
WIN_ROW     = PatternFill("solid", fgColor="FFFDE7")   # winner row accent
THIN        = Side(style="thin", color="BDBDBD")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ─────────────────────── public API ──────────────────────────────────────────

def export_generated_games(tickets: List[Ticket], dirpath: str) -> str:
    """Export freshly generated tickets to Excel. Returns file path."""
    os.makedirs(dirpath, exist_ok=True)
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    cfg = LOTTERY_CONFIGS[tickets[0].lottery_type]
    fname = f"{cfg.display_name.replace(' ', '_').replace('+', 'Mais')}_{ts}.xlsx"
    path  = os.path.join(dirpath, fname)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # group by lottery type (in case mixed batch ever added)
    by_type: dict[LotteryType, List[Ticket]] = {}
    for t in tickets:
        by_type.setdefault(t.lottery_type, []).append(t)

    for lt, lt_tickets in by_type.items():
        lt_cfg = LOTTERY_CONFIGS[lt]
        ws = wb.create_sheet(lt_cfg.display_name[:31])
        _write_games_sheet(ws, lt_tickets, lt)

    _write_summary_sheet(wb, tickets)
    wb.save(path)
    return path


def export_check_results(
    results: List[CheckResult],
    draw: DrawResult,
    dirpath: str,
    ticket_price: float,
) -> str:
    """Export check results with matched highlights and financial summary."""
    os.makedirs(dirpath, exist_ok=True)
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    lt  = draw.lottery_type
    cfg = LOTTERY_CONFIGS[lt]
    contest = draw.contest_number or "s-n"
    fname = f"Resultado_{cfg.display_name.replace(' ', '_')}_C{contest}_{ts}.xlsx"
    path  = os.path.join(dirpath, fname)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_res = wb.create_sheet("Resultado")
    _write_result_sheet(ws_res, results, draw, lt)

    ws_fin = wb.create_sheet("Resumo Financeiro")
    _write_financial_sheet(ws_fin, results, draw, lt, ticket_price)

    wb.save(path)
    return path


# ─────────────────────── games sheet ─────────────────────────────────────────

def _write_games_sheet(ws: Worksheet, tickets: List[Ticket], lt: LotteryType):
    cfg   = LOTTERY_CONFIGS[lt]
    color = BRAND.get(lt, "1A7F44")
    light = LIGHT.get(lt, "E8F5E9")
    hfill = PatternFill("solid", fgColor=color)
    nfill = PatternFill("solid", fgColor=light)
    hfont = Font(bold=True, color="FFFFFF", size=11)
    tfont = Font(bold=True, color="FFFFFF", size=13)

    is_loto = (lt == LotteryType.LOTOMANIA)

    # ── title row
    title_cols = 10 if is_loto else (5 + len(tickets[0].numbers))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(title_cols, 8))
    tc = ws.cell(1, 1, f"{cfg.display_name} – {len(tickets)} Jogo(s) – {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    tc.font = tfont
    tc.fill = hfill
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── header row
    if is_loto:
        headers = ["#", "ID", "Concurso", "50 Dezenas (01–100)", "Acertos", "Prêmio", "Data"]
    else:
        num_cols = [f"N{i+1}" for i in range(len(tickets[0].numbers))]
        extra_h  = [cfg.extra_name] if cfg.extra_name else []
        headers  = ["#", "ID", "Concurso"] + num_cols + extra_h + ["Acertos", "Prêmio", "Data"]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(2, c, h)
        cell.font      = hfont
        cell.fill      = hfill
        cell.alignment = Alignment(horizontal="center")
        cell.border    = BORDER
    ws.row_dimensions[2].height = 22

    # ── data rows
    for r, ticket in enumerate(tickets, 3):
        alt = PatternFill("solid", fgColor="F5F5F5") if r % 2 == 0 else None

        def _cell(col: int, value, fill=None, font=None, align="center"):
            c = ws.cell(r, col, value)
            c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            c.border    = BORDER
            if fill: c.fill = fill
            if font: c.font = font
            return c

        _cell(1, r - 2, alt)
        _cell(2, ticket.id, alt)
        _cell(3, "", alt)   # Concurso – blank for user

        if is_loto:
            # All 50 numbers in one cell, formatted in rows of 10
            lines = []
            for i in range(0, 50, 10):
                chunk = ticket.numbers[i:i+10]
                lines.append("  ".join(f"{n:02d}" for n in chunk))
            _cell(4, "\n".join(lines), nfill, align="left")
            ws.row_dimensions[r].height = 75
            _cell(5, "", alt)   # Acertos
            _cell(6, "", alt)   # Prêmio
            _cell(7, ticket.created_at, alt)
        else:
            col = 4
            for num in ticket.numbers:
                _cell(col, num, nfill, Font(bold=True, color="333333"))
                col += 1
            if cfg.extra_name:
                _cell(col, _fmt_extra(ticket, lt), alt)
                col += 1
            _cell(col,     "", alt)   # Acertos
            _cell(col + 1, "", alt)   # Prêmio
            _cell(col + 2, ticket.created_at, alt)

    # ── column widths
    _auto_width(ws)
    ws.freeze_panes = "A3"


# ─────────────────────── result sheet ────────────────────────────────────────

def _write_result_sheet(ws: Worksheet, results: List[CheckResult],
                        draw: DrawResult, lt: LotteryType):
    cfg    = LOTTERY_CONFIGS[lt]
    color  = BRAND.get(lt, "1A7F44")
    hfill  = PatternFill("solid", fgColor=color)
    hfont  = Font(bold=True, color="FFFFFF", size=11)
    tfont  = Font(bold=True, color="FFFFFF", size=13)
    is_loto = (lt == LotteryType.LOTOMANIA)

    draw_set  = set(draw.numbers)
    draw_set2 = set(draw.numbers2) if draw.numbers2 else set()

    # title
    n_cols = 10 if is_loto else (5 + (results[0].ticket.numbers.__len__() if results else 10))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(n_cols, 9))
    contest_str = f"Concurso {draw.contest_number}" if draw.contest_number else ""
    date_str    = f"  {draw.draw_date}" if draw.draw_date else ""
    tc = ws.cell(1, 1, f"{cfg.display_name} – Resultado {contest_str}{date_str}")
    tc.font = tfont; tc.fill = hfill
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # drawn numbers info row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(n_cols, 9))
    draw_str = "  ".join(f"{n:02d}" for n in sorted(draw.numbers))
    info = f"Números sorteados: {draw_str}"
    if draw.numbers2:
        info += "  |  2º: " + "  ".join(f"{n:02d}" for n in sorted(draw.numbers2))
    if draw.extra is not None:
        info += f"  |  Extra: {_fmt_draw_extra(draw, lt)}"
    ic = ws.cell(2, 1, info)
    ic.font = Font(bold=True, size=11, color="333333")
    ic.fill = PatternFill("solid", fgColor="FFF9C4")
    ic.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    # headers
    if is_loto:
        headers = ["#", "ID", "Rótulo", "50 Dezenas", "Acertos", "Prêmio", "Observação"]
    else:
        sample = results[0].ticket.numbers if results else []
        num_cols = [f"N{i+1}" for i in range(len(sample))]
        extra_h  = [cfg.extra_name] if cfg.extra_name else []
        d2_h     = ["Acertos 2º"] if cfg.has_second_draw else []
        headers  = ["#", "ID", "Rótulo"] + num_cols + extra_h + ["Acertos"] + d2_h + ["Prêmio"]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.font = hfont; cell.fill = hfill
        cell.alignment = Alignment(horizontal="center"); cell.border = BORDER
    ws.row_dimensions[3].height = 22

    # data
    for r, res in enumerate(results, 4):
        ticket = res.ticket
        won    = res.is_winner
        row_fill = WIN_ROW if won else None

        def _cell(col: int, value, fill=None, font=None, align="center"):
            c = ws.cell(r, col, value)
            c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            c.border = BORDER
            if fill: c.fill = fill
            elif row_fill: c.fill = row_fill
            if font: c.font = font
            return c

        _cell(1, r - 3)
        _cell(2, ticket.id)
        _cell(3, ticket.label or "—", align="left")

        if is_loto:
            lines = []
            for i in range(0, 50, 10):
                chunk = ticket.numbers[i:i+10]
                parts = []
                for n in chunk:
                    if n in draw_set:
                        parts.append(f"[{n:02d}]")   # bracket = matched
                    else:
                        parts.append(f" {n:02d} ")
                lines.append("".join(parts))
            _cell(4, "\n".join(lines), align="left")
            ws.row_dimensions[r].height = 75
            _cell(5, res.matches)
            _cell(6, res.best_prize.name if res.best_prize else "—",
                  font=Font(bold=True, color="1A7F44") if won else None)
            obs = "🏆 PREMIADO!" if won else ""
            _cell(7, obs, font=Font(bold=True, color="D4820A") if won else None)
        else:
            col = 4
            for num in ticket.numbers:
                matched = num in draw_set
                nf = MATCH_FILL if matched else PatternFill("solid", fgColor=LIGHT.get(lt, "E8F5E9"))
                nft = MATCH_FONT if matched else Font(bold=True, color="555555")
                _cell(col, num, nf, nft)
                col += 1

            if cfg.extra_name:
                extra_ok = res.extra_matches > 0
                ef = PatternFill("solid", fgColor="27AE60") if extra_ok else None
                eft = Font(bold=True, color="FFFFFF") if extra_ok else None
                _cell(col, _fmt_extra(ticket, lt), ef, eft)
                col += 1

            prize_str = res.prize_tier.name if res.prize_tier else "—"
            prize_font = Font(bold=True, color="1A7F44") if res.prize_tier else None
            acertos_font = Font(bold=True, color="27AE60") if res.matches > 0 else None

            _cell(col, res.matches, font=acertos_font); col += 1

            if cfg.has_second_draw:
                a2f = Font(bold=True, color="27AE60") if res.matches2 > 0 else None
                _cell(col, res.matches2, font=a2f); col += 1
                if res.prize_tier2:
                    prize_str += f" | 2º: {res.prize_tier2.name}"

            _cell(col, prize_str, font=prize_font)

    _auto_width(ws)
    ws.freeze_panes = "A4"


# ─────────────────────── financial sheet ─────────────────────────────────────

def _write_financial_sheet(ws: Worksheet, results: List[CheckResult],
                           draw: DrawResult, lt: LotteryType,
                           ticket_price: float):
    cfg   = LOTTERY_CONFIGS[lt]
    color = BRAND.get(lt, "1A7F44")

    hfill = PatternFill("solid", fgColor=color)
    hfont = Font(bold=True, color="FFFFFF", size=12)
    tfont = Font(bold=True, color="FFFFFF", size=14)
    bfont = Font(bold=True, size=11)

    # title
    ws.merge_cells("A1:E1")
    tc = ws.cell(1, 1, f"Resumo Financeiro – {cfg.display_name}")
    tc.font = tfont; tc.fill = hfill
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # helpers
    def _label(row, col, text):
        c = ws.cell(row, col, text)
        c.font = bfont; c.border = BORDER
        c.alignment = Alignment(horizontal="left", vertical="center")

    def _value(row, col, text, color_str="333333", bold=True):
        c = ws.cell(row, col, text)
        c.font = Font(bold=bold, size=11, color=color_str)
        c.border = BORDER
        c.alignment = Alignment(horizontal="right", vertical="center")

    def _sect(row, text):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row, 1, text)
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="607D8B")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22

    # ── Investimento
    n_games      = len(results)
    total_spent  = sum(r.ticket.cost() for r in results)

    _sect(2, "💰  Investimento")
    _label(3, 1, "Jogos realizados");      _value(3, 2, n_games)
    _label(4, 1, "Preço da aposta simples"); _value(4, 2, f"R$ {ticket_price:.2f}")
    _label(5, 1, "Total investido");       _value(5, 2, f"R$ {total_spent:.2f}", "B71C1C")

    # ── Resultado
    winners    = [r for r in results if r.is_winner]
    n_win      = len(winners)
    n_loss     = n_games - n_win

    fixed_total = 0.0
    var_prizes  = []
    for r in winners:
        prize = r.best_prize
        if prize:
            if prize.prize_type == "fixed":
                fixed_total += prize.fixed_value
            else:
                var_prizes.append(prize.name)
        # Dupla-Sena: also check 2nd draw
        if r.prize_tier2 and r.prize_tier2 != r.prize_tier:
            if r.prize_tier2.prize_type == "fixed":
                fixed_total += r.prize_tier2.fixed_value
            elif r.prize_tier2.name not in var_prizes:
                var_prizes.append(r.prize_tier2.name)

    _sect(7, "🏆  Resultado dos Jogos")
    _label(8, 1,  "Jogos premiados");  _value(8, 2, n_win,  "1A7F44")
    _label(9, 1,  "Jogos sem prêmio"); _value(9, 2, n_loss, "B71C1C")
    _label(10, 1, "Prêmios fixos (R$)"); _value(10, 2, f"R$ {fixed_total:.2f}", "1A7F44")

    if var_prizes:
        unique_var = list(dict.fromkeys(var_prizes))
        _label(11, 1, "Prêmios variáveis"); _value(11, 2, " | ".join(unique_var), "D4820A")
        next_row = 12
    else:
        next_row = 11

    # ── Balanço
    _sect(next_row, "📊  Balanço Final (prêmios fixos)")
    r1 = next_row + 1
    net  = fixed_total - total_spent
    roi  = (fixed_total / total_spent * 100) if total_spent else 0
    net_color  = "1A7F44" if net >= 0 else "B71C1C"
    roi_color  = "1A7F44" if roi >= 100 else "B71C1C"

    _label(r1,     1, "Total investido");      _value(r1,     2, f"R$ {total_spent:.2f}", "B71C1C")
    _label(r1 + 1, 1, "Retorno fixo total");   _value(r1 + 1, 2, f"R$ {fixed_total:.2f}", "1A7F44")
    _label(r1 + 2, 1, "Saldo líquido");        _value(r1 + 2, 2, f"R$ {net:+.2f}", net_color)
    _label(r1 + 3, 1, "ROI (retorno/custo)");  _value(r1 + 3, 2, f"{roi:.1f}%", roi_color)

    if var_prizes:
        note_row = r1 + 5
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
        nc = ws.cell(note_row, 1,
                     "* Prêmios variáveis (jackpot/rateio) não estão incluídos no saldo. "
                     "Consulte a Caixa para o valor exato.")
        nc.font = Font(italic=True, color="757575", size=9)
        nc.alignment = Alignment(wrap_text=True)

    # widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    for col in "CDE":
        ws.column_dimensions[col].width = 14


# ─────────────────────── summary sheet (for generation) ──────────────────────

def _write_summary_sheet(wb: openpyxl.Workbook, tickets: List[Ticket]):
    ws = wb.create_sheet("Resumo", 0)
    hfill = PatternFill("solid", fgColor="37474F")
    hfont = Font(bold=True, color="FFFFFF", size=14)

    ws.merge_cells("A1:D1")
    tc = ws.cell(1, 1, f"Jogos Gerados – {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    tc.font = hfont; tc.fill = hfill
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.cell(3, 1, "Tipo de Jogo").font  = Font(bold=True)
    ws.cell(3, 2, "Qtd. Jogos").font    = Font(bold=True)
    ws.cell(3, 3, "Preço base").font    = Font(bold=True)
    ws.cell(3, 4, "Total Investido").font = Font(bold=True)

    by_type: dict[LotteryType, list] = {}
    for t in tickets:
        by_type.setdefault(t.lottery_type, []).append(t)

    total_inv = 0.0
    for r, (lt, lt_tickets) in enumerate(by_type.items(), 4):
        cfg     = LOTTERY_CONFIGS[lt]
        inv     = sum(t.cost() for t in lt_tickets)
        total_inv += inv
        color   = BRAND.get(lt, "333333")
        ws.cell(r, 1, f"{cfg.emoji} {cfg.display_name}").font = Font(color=color, bold=True)
        ws.cell(r, 2, len(lt_tickets))
        ws.cell(r, 3, f"R$ {cfg.ticket_price:.2f}")
        ws.cell(r, 4, f"R$ {inv:.2f}")

    total_row = 4 + len(by_type)
    ws.cell(total_row, 1, "TOTAL").font = Font(bold=True)
    ws.cell(total_row, 2, len(tickets)).font = Font(bold=True)
    ws.cell(total_row, 4, f"R$ {total_inv:.2f}").font = Font(bold=True, color="B71C1C")

    for col in "ABCD":
        ws.column_dimensions[col].width = 24


# ─────────────────────── helpers ─────────────────────────────────────────────

def _auto_width(ws: Worksheet):
    for col in ws.columns:
        best = 0
        for cell in col:
            if cell.value:
                val = str(cell.value)
                best = max(best, max(len(ln) for ln in val.split("\n")))
        ltr = get_column_letter(col[0].column)
        ws.column_dimensions[ltr].width = min(max(best + 2, 8), 60)


def _fmt_extra(ticket: Ticket, lt: LotteryType) -> str:
    if ticket.extra is None:
        return "—"
    if lt == LotteryType.TIMEMANIA:
        return f"{TIMEMANIA_TEAMS.get(ticket.extra, ticket.extra)}"
    if lt == LotteryType.DIA_DE_SORTE:
        return f"{DIA_DE_SORTE_MONTHS.get(ticket.extra, ticket.extra)}"
    if lt == LotteryType.MAIS_MILIONARIA:
        return f"{ticket.extra[0]} e {ticket.extra[1]}"
    return str(ticket.extra)


def _fmt_draw_extra(draw: DrawResult, lt: LotteryType) -> str:
    if draw.extra is None:
        return "—"
    if lt == LotteryType.TIMEMANIA:
        return str(TIMEMANIA_TEAMS.get(draw.extra, draw.extra))
    if lt == LotteryType.DIA_DE_SORTE:
        return str(DIA_DE_SORTE_MONTHS.get(draw.extra, draw.extra))
    if lt == LotteryType.MAIS_MILIONARIA:
        ext = draw.extra
        if isinstance(ext, list):
            return f"{ext[0]} e {ext[1]}"
    return str(draw.extra)
